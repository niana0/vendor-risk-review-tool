"""
Questionnaire Mapper - RAG pipeline: embed chunks, retrieve relevant ones per question, LLM answers
"""
from typing import Dict, List, Any
import json
import os
import numpy as np
import openpyxl
from openai import OpenAI


class QuestionnaireMapper:

    EMBEDDING_MODEL = "text-embedding-3-large"
    COMPLETION_MODEL = "gpt-4o"
    TOP_K_CHUNKS = 8        # number of relevant chunks to retrieve per question
    QUESTIONS_PER_BATCH = 5 # questions per LLM call

    def __init__(self):
        api_key = os.getenv("OPENAI_API_KEY")
        self.client = OpenAI(api_key=api_key) if api_key else None
        self.question_mappings = []

    # ------------------------------------------------------------------
    # Embedding helpers
    # ------------------------------------------------------------------

    def _get_embeddings(self, texts: List[str]) -> np.ndarray:
        """Batch embed a list of texts. Returns (N, D) numpy array."""
        if not texts:
            return np.array([])
        response = self.client.embeddings.create(
            model=self.EMBEDDING_MODEL,
            input=texts
        )
        ordered = sorted(response.data, key=lambda x: x.index)
        return np.array([item.embedding for item in ordered])

    def _cosine_similarity_matrix(self, a: np.ndarray, b: np.ndarray) -> np.ndarray:
        """Returns (len(a), len(b)) similarity matrix."""
        a_norm = a / np.linalg.norm(a, axis=1, keepdims=True)
        b_norm = b / np.linalg.norm(b, axis=1, keepdims=True)
        return np.dot(a_norm, b_norm.T)

    # ------------------------------------------------------------------
    # Questionnaire loading
    # ------------------------------------------------------------------

    def load_questionnaire(self, excel_path: str) -> List[Dict[str, Any]]:
        questions = []
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet_names = ['Questions', 'Questionnaire', 'Assessment', workbook.sheetnames[0]]
            sheet = None
            for name in sheet_names:
                if name in workbook.sheetnames:
                    sheet = workbook[name]
                    break
            if not sheet:
                return questions

            # Detect header row to find which column contains the question text
            header_row = [cell.value for cell in sheet[1]]
            question_col_idx = None
            question_keywords = ['question', 'control', 'requirement', 'description', 'item']
            for i, header in enumerate(header_row):
                if header and any(kw in str(header).lower() for kw in question_keywords):
                    question_col_idx = i
                    break

            q_num = 1
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue

                question_text = None
                if question_col_idx is not None:
                    # Use the detected question column
                    cell = row[question_col_idx] if question_col_idx < len(row) else None
                    if cell and isinstance(cell, str) and len(cell.strip()) > 5:
                        question_text = cell.strip()
                else:
                    # Fall back: use the longest string cell in the row as the question
                    candidates = [
                        c.strip() for c in row
                        if c and isinstance(c, str) and len(c.strip()) > 20
                    ]
                    if candidates:
                        question_text = max(candidates, key=len)

                if question_text:
                    questions.append({
                        "id": f"Q{q_num}",
                        "question": question_text,
                        "category": "General",
                        "row_num": q_num + 1
                    })
                    q_num += 1
        except Exception as e:
            print(f"Error loading questionnaire: {e}")
        return questions

    # ------------------------------------------------------------------
    # RAG pipeline
    # ------------------------------------------------------------------

    def map_evidence_to_questions(
        self,
        questions: List[Dict[str, Any]],
        evidence_library: List[Dict[str, Any]],
        threshold: float = 0.2
    ) -> List[Dict[str, Any]]:
        self.question_mappings = []

        if not self.client:
            for q in questions:
                self.question_mappings.append(self._empty_mapping(q, "OpenAI API key not configured"))
            return self.question_mappings

        # Build flat list of chunks with their source metadata
        chunks = [
            {
                "text": ev.get("text") or ev.get("context", ""),
                "source": ev.get("source", "Unknown")
            }
            for ev in evidence_library
            if ev.get("text") or ev.get("context")
        ]

        if not chunks:
            for q in questions:
                self.question_mappings.append(self._empty_mapping(q, "No document content extracted"))
            return self.question_mappings

        chunk_texts = [c["text"] for c in chunks]
        question_texts = [q["question"] for q in questions]

        # Embed everything in 2 API calls
        print(f"Embedding {len(chunk_texts)} document chunks...")
        chunk_embeddings = self._get_embeddings(chunk_texts)

        print(f"Embedding {len(question_texts)} questions...")
        question_embeddings = self._get_embeddings(question_texts)

        # Compute full similarity matrix: (num_questions, num_chunks)
        similarity_matrix = self._cosine_similarity_matrix(question_embeddings, chunk_embeddings)

        # For each question, retrieve top-K chunks
        question_chunks: List[List[Dict]] = []
        for i in range(len(questions)):
            scores = similarity_matrix[i]
            top_indices = np.argsort(scores)[::-1][:self.TOP_K_CHUNKS]
            retrieved = [
                {**chunks[j], "score": float(scores[j])}
                for j in top_indices
                if float(scores[j]) > 0.1  # discard completely unrelated chunks
            ]
            question_chunks.append(retrieved)

        # Send to LLM in batches
        for i in range(0, len(questions), self.QUESTIONS_PER_BATCH):
            batch_questions = questions[i:i + self.QUESTIONS_PER_BATCH]
            batch_chunks = question_chunks[i:i + self.QUESTIONS_PER_BATCH]
            mappings = self._answer_batch(batch_questions, batch_chunks)
            self.question_mappings.extend(mappings)

        return self.question_mappings

    def _answer_batch(
        self,
        questions: List[Dict[str, Any]],
        question_chunks: List[List[Dict]]
    ) -> List[Dict[str, Any]]:
        """Send a batch of questions, each with their own retrieved chunks, to the LLM."""

        # Build prompt sections per question
        sections = []
        for i, (q, chunks) in enumerate(zip(questions, question_chunks)):
            chunk_text = ""
            for chunk in chunks:
                chunk_text += f"\n[Source: {chunk['source']}]\n{chunk['text']}\n"

            if not chunk_text.strip():
                chunk_text = "No relevant excerpts found."

            sections.append(
                f"--- Question {i + 1} [ID: {q['id']}] ---\n"
                f"{q['question']}\n\n"
                f"Relevant excerpts from vendor documentation:\n{chunk_text}"
            )

        prompt = (
            "You are a security analyst completing a vendor security questionnaire.\n\n"
            "For each question below, relevant excerpts from the vendor's documentation have been retrieved. "
            "Read each question and its excerpts carefully, then:\n"
            "1. Start the answer with 'Yes', 'No', or 'Partially' based on whether the documentation addresses the question\n"
            "2. Follow with a specific explanation based only on the provided excerpts\n"
            "3. If the excerpts partially address the question, start with 'Partially' and state what is covered and what is missing\n"
            "4. If the excerpts do not address the question at all, set answer to "
            "'Not addressed in the provided documentation' and confidence to 'NOT_FOUND'\n"
            "5. Always include the source (document name and page) where the answer was found\n\n"
            + "\n\n".join(sections) +
            "\n\nRespond with a JSON object in this exact format:\n"
            '{"answers": [\n'
            '  {"question_id": "Q2", "answer": "...", "source": "filename.pdf, Page 3", "confidence": "HIGH"},\n'
            '  ...\n'
            ']}\n\n'
            "Confidence: HIGH (directly answered), MEDIUM (partially addressed), "
            "LOW (indirectly mentioned), NOT_FOUND (not in documentation)"
        )

        try:
            response = self.client.chat.completions.create(
                model=self.COMPLETION_MODEL,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=4000,
                temperature=0,
                response_format={"type": "json_object"}
            )
            result = json.loads(response.choices[0].message.content)
            answers = result.get("answers", [])
        except Exception as e:
            print(f"LLM batch failed: {e}")
            answers = []

        answer_map = {a["question_id"]: a for a in answers}
        mappings = []

        for q, chunks in zip(questions, question_chunks):
            data = answer_map.get(q["id"], {})
            answer = data.get("answer", "Not addressed in the provided documentation")
            confidence = data.get("confidence", "NOT_FOUND")
            source = data.get("source", "")

            evidence = [
                {"source": c["source"], "evidence_text": c["text"], "similarity_score": c["score"]}
                for c in chunks
            ] if chunks else []

            gaps = [] if confidence in ("HIGH", "MEDIUM") else ["Not addressed in the provided documentation"]

            mappings.append({
                "question_id": q["id"],
                "question": q["question"],
                "category": q.get("category", "General"),
                "answer": answer,
                "evidence": evidence,
                "confidence": confidence,
                "gaps": gaps
            })

        return mappings

    def _empty_mapping(self, question: Dict[str, Any], reason: str) -> Dict[str, Any]:
        return {
            "question_id": question["id"],
            "question": question["question"],
            "category": question.get("category", "General"),
            "answer": reason,
            "evidence": [],
            "confidence": "NOT_FOUND",
            "gaps": [reason]
        }

    # ------------------------------------------------------------------
    # Output
    # ------------------------------------------------------------------

    def to_json(self, output_path: str = "questionnaire_mapping.json"):
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.question_mappings, f, indent=2, ensure_ascii=False)
        return output_path

    def to_excel(self, output_path: str = "completed_questionnaire.xlsx"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Completed Assessment"

        headers = ["Question ID", "Category", "Question", "Answer", "Evidence References", "Confidence", "Gaps/Follow-ups"]
        sheet.append(headers)

        for mapping in self.question_mappings:
            evidence_refs = "; ".join([
                e["source"] for e in mapping["evidence"][:3] if e.get("source")
            ])
            gaps = "; ".join(mapping["gaps"]) if mapping["gaps"] else "None"

            sheet.append([
                mapping["question_id"],
                mapping["category"],
                mapping["question"],
                mapping["answer"],
                evidence_refs,
                mapping["confidence"],
                gaps
            ])

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 80)

        workbook.save(output_path)
        return output_path
